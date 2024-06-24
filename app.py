from flask import Flask,render_template,request
from openpyxl import load_workbook
from datetime import datetime

wb=load_workbook("Robotics_Registration.xlsx")
ws=wb.active
next_row=ws.max_row+1


app = Flask(__name__)

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/submit', methods =['POST'])
def submit():
    submissionDate = datetime.today().date()
    studentFirstName = request.form.get('studentFirstName')
    studentMiddleName = request.form['studentMiddleName']
    studentLastName = request.form['studentLastName']
    dateOfBirth = request.form['dateOfBirth']
    gender = request.form['gender']
    phoneNumber = request.form['phoneNumber']
    email = request.form['email']
    schoolName = request.form['schoolName']
    grade = request.form['grade']
    parentFirstName = request.form['parentFirstName']
    parentMiddleName = request.form['parentMiddleName']
    parentLastName = request.form['parentLastName']
    street = request.form['street']
    city = request.form['city']
    postalCode = request.form['postalCode']
    secondPhone = request.form['secondPhone']
    remarks = request.form['remarks']

    ws[f"A{next_row}"] = submissionDate
    ws[f"B{next_row}"] = f"{studentFirstName} {studentMiddleName} {studentLastName}"
    ws[f"C{next_row}"] = dateOfBirth
    ws[f"E{next_row}"] = gender
    ws[f"F{next_row}"] = phoneNumber
    ws[f"G{next_row}"] = email
    ws[f"H{next_row}"] = schoolName
    ws[f"I{next_row}"] = grade
    ws[f"J{next_row}"] = f"{parentFirstName} {parentMiddleName} {parentLastName}"
    ws[f"K{next_row}"] = f"{street}, {city}, {postalCode}"
    ws[f"L{next_row}"] = secondPhone
    ws[f"M{next_row}"] = remarks

    wb.save("Robotics_Registration.xlsx")

    print( f"""{submissionDate},{studentFirstName} {studentMiddleName} {studentLastName} {
dateOfBirth} {gender} {phoneNumber} {email} {schoolName} {grade}
{parentFirstName} {parentMiddleName} {parentLastName} {street}, {city}, {postalCode}
{secondPhone} {remarks}""")

    return f"""{submissionDate},{studentFirstName} {studentMiddleName} {studentLastName} {
dateOfBirth} {gender} {phoneNumber} {email} {schoolName} {grade}
{parentFirstName} {parentMiddleName} {parentLastName} {street}, {city}, {postalCode}
{secondPhone} {remarks}"""
    
if __name__=='__main__':
    app.run(debug=True)
